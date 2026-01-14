"""Export all tables to a single Excel file with individual sheets."""

import sqlite3
from pathlib import Path
from db_new import get_conn

try:
    import openpyxl
    from openpyxl import Workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def export_to_excel(output_file: str = "company_data_export.xlsx") -> None:
    """Export description_industry to separate file; other tables to single file."""

    if not HAS_OPENPYXL:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        print("Falling back to CSV export...")
        export_to_csv()
        return

    # Prepare output directory
    output_dir = Path("exports")
    output_dir.mkdir(exist_ok=True)

    conn = get_conn()

    # --- SEPARATE FILE 1: description_industry ---
    wb_desc = Workbook()
    ws_desc = wb_desc.active
    ws_desc.title = "description_industry"
    desc_cols = [
        "domain",
        "sub_industry",
        "industry",
        "sector",
        "sic_code",
        "sic_description",
    ]
    ws_desc.append(desc_cols)
    desc_rows = conn.execute(
        "SELECT domain, sub_industry, industry, sector, sic_code, sic_text FROM description_industry ORDER BY domain"
    ).fetchall()
    for r in desc_rows:
        ws_desc.append(
            [
                r["domain"],
                r["sub_industry"],
                r["industry"],
                r["sector"],
                r["sic_code"],
                r["sic_text"],
            ]
        )
    desc_path = output_dir / "description_industry.xlsx"
    wb_desc.save(desc_path.as_posix())
    print(f"✓ description_industry.xlsx: {len(desc_rows)} rows -> {desc_path}")

    # --- COMBINED FILE 2: All other tables ---
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # 1. Contact Information sheet
    ws_contact = wb.create_sheet(title="contact_information")
    contact_cols = [
        "domain",
        "text",
        "company_name",
        "full_address",
        "phone",
        "sales_phone",
        "fax",
        "mobile",
        "other_numbers",
        "email",
        "hours_of_operation",
        "hq_indicator",
    ]
    ws_contact.append(contact_cols)
    contact_rows = conn.execute(
        f"SELECT {', '.join(contact_cols)} FROM contact_information"
    ).fetchall()
    for r in contact_rows:
        ws_contact.append([r[c] for c in contact_cols])
    print(f"✓ contact_information: {len(contact_rows)} rows")

    # 2. Company Information sheet
    ws_company = wb.create_sheet(title="company_information")
    company_cols = ["domain", "company_name", "acronym", "logo_url"]
    ws_company.append(company_cols)
    company_rows = conn.execute(
        f"SELECT {', '.join(company_cols)} FROM company_information"
    ).fetchall()
    for r in company_rows:
        ws_company.append([r[c] for c in company_cols])
    print(f"✓ company_information: {len(company_rows)} rows")

    # 3. Social Media sheet
    ws_social = wb.create_sheet(title="social_media")
    social_cols = [
        "domain",
        "linkedin",
        "facebook",
        "x",
        "instagram",
        "youtube",
        "blog",
        "articles",
    ]
    ws_social.append(social_cols)
    social_rows = conn.execute(
        f"SELECT {', '.join(social_cols)} FROM social_media"
    ).fetchall()
    for r in social_rows:
        ws_social.append([r[c] for c in social_cols])
    print(f"✓ social_media: {len(social_rows)} rows")

    # 4. People Information sheet
    ws_people = wb.create_sheet(title="people_information")
    people_cols = ["domain", "people_name", "people_title", "people_email", "url"]
    ws_people.append(people_cols)
    people_rows = conn.execute(
        f"SELECT {', '.join(people_cols)} FROM people_information"
    ).fetchall()
    for r in people_rows:
        ws_people.append([r[c] for c in people_cols])
    print(f"✓ people_information: {len(people_rows)} rows")

    # 5. Certifications sheet
    ws_certs = wb.create_sheet(title="certifications")
    cert_cols = ["domain", "certification"]
    ws_certs.append(cert_cols)
    cert_rows = conn.execute(
        f"SELECT {', '.join(cert_cols)} FROM certifications"
    ).fetchall()
    for r in cert_rows:
        ws_certs.append([r[c] for c in cert_cols])
    print(f"✓ certifications: {len(cert_rows)} rows")

    # 6. Services sheet
    ws_services = wb.create_sheet(title="services")
    service_cols = ["domain", "item_name", "item_type"]
    ws_services.append(service_cols)
    service_rows = conn.execute(
        f"SELECT {', '.join(service_cols)} FROM services"
    ).fetchall()
    for r in service_rows:
        ws_services.append([r[c] for c in service_cols])
    print(f"✓ services: {len(service_rows)} rows")

    # Save combined workbook
    export_path = output_dir / output_file
    wb.save(export_path.as_posix())
    conn.close()

    total = (
        len(contact_rows)
        + len(company_rows)
        + len(social_rows)
        + len(people_rows)
        + len(cert_rows)
        + len(service_rows)
    )
    print(f"\n✓ company_data_export.xlsx: {total} rows -> {export_path}")


def export_to_csv() -> None:
    """CSV fallback: one CSV per table."""
    import csv

    conn = get_conn()
    output_dir = Path("exports")
    output_dir.mkdir(exist_ok=True)

    tables = {
        "description_industry": [
            "domain",
            "sub_industry",
            "industry",
            "sector",
            "sic_code",
            "sic_text",
        ],
        "contact_information": [
            "domain",
            "text",
            "company_name",
            "full_address",
            "phone",
            "sales_phone",
            "fax",
            "mobile",
            "other_numbers",
            "email",
            "hours_of_operation",
            "hq_indicator",
        ],
        "company_information": ["domain", "company_name", "acronym", "logo_url"],
        "social_media": [
            "domain",
            "linkedin",
            "facebook",
            "x",
            "instagram",
            "youtube",
            "blog",
            "articles",
        ],
        "people_information": [
            "domain",
            "people_name",
            "people_title",
            "people_email",
            "url",
        ],
        "certifications": ["domain", "certification"],
        "services": ["domain", "item_name", "item_type"],
    }

    for table_name, columns in tables.items():
        rows = conn.execute(f"SELECT {', '.join(columns)} FROM {table_name}").fetchall()
        csv_path = output_dir / f"{table_name}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for r in rows:
                writer.writerow([r[c] for c in columns])
        print(f"✓ {table_name}.csv: {len(rows)} rows")

    conn.close()
    print(f"\n✓ CSV exports complete in: {output_dir}/")


if __name__ == "__main__":
    print("Exporting company data...\n")
    export_to_excel()
